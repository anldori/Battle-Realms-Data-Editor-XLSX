"""
brde.core - Read/write engine for Battle Realms.xlsx

Reading: openpyxl in read-only mode, which is fast and only needs the values.
Writing: NOT openpyxl. Instead the XML inside the .xlsx archive is patched in
         place, so every part of the file that was not edited stays byte-for-byte
         identical and the game keeps reading it exactly as before.
"""
from __future__ import annotations

import os
import re
import shutil
import zipfile
from datetime import datetime

import openpyxl

from . import dat, schema

NS_MAIN = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

# The game's old binary format. Opened read-only: see BRWorkbook.read_only.
DAT_SUFFIX = '.dat'


# ------------------------------------------------------------------ utils
def col_letter(idx0: int) -> str:
    """0 -> A, 25 -> Z, 26 -> AA"""
    s = ''
    n = idx0 + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_index(letter: str) -> int:
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def xml_escape(s: str) -> str:
    return (str(s).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;'))


def _cleanup_old_backups(dest: str, keep_count: int):
    """Remove old .bak files, keeping the most recent keep_count backups."""
    if keep_count <= 0:
        return
    dir_path = os.path.dirname(dest) or '.'
    base_name = os.path.basename(dest)
    backups = []
    for name in os.listdir(dir_path):
        if name.startswith(base_name) and name.endswith('.bak'):
            full_path = os.path.join(dir_path, name)
            mtime = os.path.getmtime(full_path)
            backups.append((mtime, full_path))
    backups.sort(reverse=True)
    for _, path in backups[keep_count:]:
        try:
            os.remove(path)
        except OSError:
            pass


# ------------------------------------------------------------------ model
class EnumTable:
    """One Enum_* sheet: numeric code <-> description."""

    __slots__ = ('name', 'items', 'code2desc', 'codes')

    def __init__(self, name, items):
        self.name = name
        self.items = items                      # [(code, desc, group)]
        self.code2desc = {c: d for c, d, _ in items}
        self.codes = set(self.code2desc)

    def label(self, code):
        d = self.code2desc.get(code)
        return f'{code} - {d}' if d else str(code)


# -1 is the workbook's usual "none / invalid", and a few switches use it instead
# of 0. It is listed so those cells read as "-1 - None" rather than being flagged
# red as a code missing from the table. No column in the vanilla file uses -1 and
# 0 and 1 together, so the extra entry costs nothing.
BOOL_TABLE = EnumTable('@bool', [(-1, 'None (-1)', ''),
                                 (0, 'No (0)', ''), (1, 'Yes (1)', '')])


class SheetData:
    __slots__ = ('name', 'headers', 'rows', 'col_enum', 'self_enum',
                 'first_data_row', 'ncols', 'colours', 'col_colour')

    def __init__(self, name, headers, rows):
        self.name = name
        self.headers = headers
        self.rows = rows                        # list[list], may contain None
        self.col_enum = {}                      # col_index -> enum name / '@bool'
        self.self_enum = None
        self.colours = []                       # [schema.ColourGroup]
        self.col_colour = {}                    # col_index -> its ColourGroup
        self.first_data_row = 2                 # row 1 is the header (1-based)
        self.ncols = len(headers)


class BRWorkbook:
    """In-memory copy of the whole workbook plus the queue of pending edits."""

    def __init__(self, path: str, progress=None):
        self.path = path
        # A .dat is the format the workbook replaced, and it is only ever read.
        # Everything above this class works off `read_only` rather than off the
        # file name: the two grids drop ItemIsEditable, the editing actions grey
        # out, and `save()` refuses outright.
        self.read_only = path.lower().endswith(DAT_SUFFIX)
        self.enums: dict[str, EnumTable] = {}
        self.sheets: dict[str, SheetData] = {}
        self.sheet_order: list[str] = []
        self.edits: dict[tuple[str, int, int], object] = {}   # (sheet,row0,col0)->value
        self.original: dict[tuple[str, int, int], object] = {}
        # Populated by the Compare feature: (sheet, row0, col0) -> (mine, theirs)
        self.diff_cells: dict[tuple[str, int, int], tuple] = {}
        self.diff_label = ''
        self._enum2sheet = None
        self._load(progress)

    # ------------------------------------------------------------ loading
    def _load(self, progress=None):
        raw = (self._read_dat if self.read_only else self._read_xlsx)(progress)
        self._build(raw)

    def _read_dat(self, progress=None) -> dict:
        """A .dat in the same {sheet: [header row, ...]} shape as a workbook.

        The two formats hold the same data model, so the whole of `_build` -
        enum mapping, colour groups, the lot - applies to both and there is no
        second code path to keep in step.

        Two things worth knowing. The .dat names its tables in full, and four of
        those names are too long for an Excel sheet and are truncated in the
        .xlsx; the full name is what is shown here, because this really is the
        old file and renaming its tables to match a workbook that is not open
        would be a lie. And the enum sheets get two columns rather than the
        workbook's three: the .dat has no SymbolGroup, and inventing a blank one
        would put an empty column on screen that no file ever had.
        """
        d = dat.DatFile(self.path, progress=progress)
        raw = {}
        for name in d.table_order:
            raw[name] = d.tables[name].as_sheet()
        for key in d.enum_order:
            raw['Enum_' + key] = ([['Code', 'Description']]
                                  + [[c, desc]
                                     for c, desc in d.enums[key].items])
        return raw

    def _read_xlsx(self, progress=None) -> dict:
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        names = wb.sheetnames
        total = len(names)
        raw = {}
        for i, name in enumerate(names):
            if progress:
                progress(i, total, name)
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            # drop trailing blank rows
            while rows and all(v is None for v in rows[-1]):
                rows.pop()
            raw[name] = rows
        wb.close()
        return raw

    def _build(self, raw: dict):
        # 1. enum tables
        for name, rows in raw.items():
            if not name.startswith('Enum_'):
                continue
            items = []
            for r in rows[1:]:
                code = r[0] if len(r) > 0 else None
                if not isinstance(code, int) or isinstance(code, bool):
                    continue
                desc = r[1] if len(r) > 1 and r[1] is not None else ''
                grp = r[2] if len(r) > 2 and r[2] is not None else ''
                items.append((int(code), str(desc), str(grp)))
            self.enums[name[5:]] = EnumTable(name[5:], items)

        # 2. data tables
        code_sets = {k: v.codes for k, v in self.enums.items()}
        for name, rows in raw.items():
            if name.startswith('Enum_'):
                continue
            if not rows:
                self.sheets[name] = SheetData(name, [], [])
                self.sheet_order.append(name)
                continue
            headers = [('' if h is None else str(h)) for h in rows[0]]
            body = rows[1:]
            sd = SheetData(name, headers, body)
            sd.self_enum = schema.sheet_self_enum(name, code_sets)
            for ci, col in enumerate(headers):
                if not col:
                    continue
                vals, clean = [], True
                for r in body:
                    if ci >= len(r) or r[ci] is None:
                        continue
                    v = r[ci]
                    if isinstance(v, int) and not isinstance(v, bool):
                        vals.append(v)
                    else:
                        clean = False
                        break
                if clean and vals:
                    e = schema.infer_column_enum(name, col, vals, code_sets,
                                                 sd.self_enum)
                    if e:
                        sd.col_enum[ci] = e
            _attach_colours(sd)
            self.sheets[name] = sd
            self.sheet_order.append(name)

        # enum sheets are browsable/editable too
        for name, rows in raw.items():
            if not name.startswith('Enum_'):
                continue
            headers = [('' if h is None else str(h)) for h in rows[0]] if rows else []
            self.sheets[name] = SheetData(name, headers, rows[1:] if rows else [])
            self.sheet_order.append(name)

    # ------------------------------------------------------------ accessors
    def data_sheet_for_enum(self, ename: str):
        """Enum_X -> the Data_* sheet keyed by X. Used by "jump to record"."""
        if self._enum2sheet is None:
            code_sets = {k: v.codes for k, v in self.enums.items()}
            primary, fallback = {}, {}
            for name in self.sheet_order:
                if name.startswith('Enum_'):
                    continue
                nat = schema.natural_self_enum(name, code_sets)
                if nat and nat not in primary:
                    primary[nat] = name
                se = self.sheets[name].self_enum
                if se and se not in fallback:
                    fallback[se] = name
            self._enum2sheet = {**fallback, **primary}
        return self._enum2sheet.get(ename)

    def colour_group(self, sheet: str, col: int):
        """The ColourGroup this column is a channel of, or None."""
        sd = self.sheets.get(sheet)
        return sd.col_colour.get(col) if sd else None

    def colour_at(self, sheet: str, row: int, group):
        """One record's colour as (r, g, b, a) floats 0..1, or None.

        None when any of red/green/blue is blank or not a number, which is what
        a half-filled row looks like: there is no colour to show yet. A missing
        alpha column, or a blank alpha cell, reads as fully opaque.
        """
        rgb = []
        for col in (group.red, group.green, group.blue):
            x = group.decode(self.value(sheet, row, col))
            if x is None:
                return None
            rgb.append(x)
        a = 1.0
        if group.alpha is not None:
            av = group.decode(self.value(sheet, row, group.alpha))
            if av is not None:
                a = av
        return (rgb[0], rgb[1], rgb[2], a)

    def enum_for(self, sheet: str, col: int):
        sd = self.sheets.get(sheet)
        if not sd:
            return None
        e = sd.col_enum.get(col)
        if e is None:
            return None
        if e == schema.BOOL_ENUM:
            return BOOL_TABLE
        return self.enums.get(e)

    def value(self, sheet: str, row: int, col: int):
        key = (sheet, row, col)
        if key in self.edits:
            return self.edits[key]
        r = self.sheets[sheet].rows
        if row < len(r) and col < len(r[row]):
            return r[row][col]
        return None

    def set_value(self, sheet: str, row: int, col: int, value):
        key = (sheet, row, col)
        r = self.sheets[sheet].rows
        orig = r[row][col] if row < len(r) and col < len(r[row]) else None
        if key not in self.original:
            self.original[key] = orig
        if value == self.original[key]:
            self.edits.pop(key, None)
        else:
            self.edits[key] = value

    @property
    def dirty(self) -> bool:
        return bool(self.edits)

    def revert_all(self):
        self.edits.clear()
        self.original.clear()

    # ------------------------------------------------------------ saving
    def _sheet_xml_map(self, zf: zipfile.ZipFile) -> dict[str, str]:
        """Sheet name -> its XML path inside the archive."""
        wbxml = zf.read('xl/workbook.xml').decode('utf-8')
        rels = zf.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rid2t = {}
        for m in re.finditer(r'<Relationship\b([^>]*?)/?>', rels):
            attrs = dict(re.findall(r'(\w+)="([^"]*)"', m.group(1)))
            if 'Id' in attrs and 'Target' in attrs:
                rid2t[attrs['Id']] = attrs['Target']

        out = {}
        for m in re.finditer(r'<sheet\b([^>]*)/>', wbxml):
            attrs = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
            name = attrs.get('name')
            rid = attrs.get('r:id') or attrs.get('id')
            tgt = rid2t.get(rid, '')
            if not tgt:
                continue
            tgt = tgt.lstrip('/')
            path = tgt if tgt.startswith('xl/') else 'xl/' + tgt
            out[_unescape(name)] = path
        return out

    def save(self, dest: str | None = None, backup: bool = True, keep_count: int = 5) -> str:
        # Saving works by patching the cells that changed inside the source
        # archive, which a .dat is not. Writing one back is a separate problem -
        # a string field is stored inline and length-prefixed, so changing one
        # shifts every byte after it and invalidates both offset tables - and
        # until that is solved the refusal belongs here rather than in the UI,
        # where a path nobody thought of could still reach it.
        if self.read_only:
            raise RuntimeError(
                'This file was opened read-only. The .dat format can be read '
                'but not written yet.')
        dest = dest or self.path
        src = self.path
        if backup and os.path.exists(dest):
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            bak = f'{dest}.{stamp}.bak'
            shutil.copy2(dest, bak)
            _cleanup_old_backups(dest, keep_count)

        by_sheet: dict[str, dict[tuple[int, int], object]] = {}
        for (sh, r, c), v in self.edits.items():
            by_sheet.setdefault(sh, {})[(r, c)] = v
        if not by_sheet:
            if dest != src:
                shutil.copy2(src, dest)
            return dest

        with zipfile.ZipFile(src, 'r') as zf:
            names = zf.namelist()
            blobs = {n: zf.read(n) for n in names}
            infos = {n: zf.getinfo(n) for n in names}
            smap = self._sheet_xml_map(zf)

        # --- shared strings
        ss_path = 'xl/sharedStrings.xml'
        ss_xml = blobs.get(ss_path, b'').decode('utf-8') if ss_path in blobs else None
        ss_list, ss_index = _parse_shared_strings(ss_xml) if ss_xml else ([], {})
        new_strings: list[str] = []

        def sst_id(text: str) -> int:
            if text in ss_index:
                return ss_index[text]
            idx = len(ss_list) + len(new_strings)
            ss_index[text] = idx
            new_strings.append(text)
            return idx

        for sheet, cells in by_sheet.items():
            path = smap.get(sheet)
            if not path or path not in blobs:
                raise RuntimeError(f'Could not find the XML part for sheet {sheet!r}')
            xml = blobs[path].decode('utf-8')
            blobs[path] = _patch_sheet(xml, cells, sst_id).encode('utf-8')

        if new_strings and ss_xml is not None:
            blobs[ss_path] = _append_shared_strings(ss_xml, new_strings).encode('utf-8')
        elif new_strings:
            raise RuntimeError('Workbook has no sharedStrings.xml; cannot write new text')

        tmp = dest + '.tmp'
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zo:
            for n in names:
                zi = infos[n]
                new = zipfile.ZipInfo(n, date_time=zi.date_time)
                new.compress_type = zi.compress_type
                new.external_attr = zi.external_attr
                new.internal_attr = zi.internal_attr
                new.create_system = zi.create_system
                zo.writestr(new, blobs[n])
        os.replace(tmp, dest)

        # sync memory: the saved values become the new baseline
        for (sh, r, c), v in self.edits.items():
            rows = self.sheets[sh].rows
            while len(rows) <= r:
                rows.append([])
            row = rows[r]
            while len(row) <= c:
                row.append(None)
            row[c] = v
        self.edits.clear()
        self.original.clear()
        self.path = dest
        return dest


def _attach_colours(sd: SheetData):
    """Find the sheet's colour groups and work out what scale they are stored on."""
    sd.colours = schema.colour_groups(sd.headers)
    for g in sd.colours:
        vals = [r[c] for r in sd.rows for c in g.columns if c < len(r)]
        g.scale = schema.colour_scale(vals)
        for c in g.columns:
            sd.col_colour[c] = g


def _unescape(s: str) -> str:
    return (s.replace('&lt;', '<').replace('&gt;', '>')
            .replace('&quot;', '"').replace('&apos;', "'")
            .replace('&amp;', '&'))


# ---------------------------------------------------- shared strings
def _parse_shared_strings(xml: str):
    items = []
    for m in re.finditer(r'<si\b.*?</si>|<si\b[^>]*/>', xml, re.S):
        block = m.group(0)
        txt = ''.join(re.findall(r'<t\b[^>]*>(.*?)</t>', block, re.S))
        items.append(_unescape(txt))
    index = {}
    for i, t in enumerate(items):
        index.setdefault(t, i)
    return items, index


def _append_shared_strings(xml: str, new_items: list[str]) -> str:
    add = ''.join(
        '<si><t xml:space="preserve">%s</t></si>' % xml_escape(t)
        for t in new_items)
    # update count / uniqueCount
    m = re.search(r'<sst\b[^>]*>', xml)
    tag = m.group(0)

    def bump(attr, delta):
        nonlocal tag
        mm = re.search(attr + r'="(\d+)"', tag)
        if mm:
            tag = tag[:mm.start()] + f'{attr}="{int(mm.group(1)) + delta}"' + tag[mm.end():]

    bump('uniqueCount', len(new_items))
    bump('count', len(new_items))
    xml = xml[:m.start()] + tag + xml[m.end():]
    idx = xml.rfind('</sst>')
    return xml[:idx] + add + xml[idx:]


# ---------------------------------------------------- patch worksheet xml
_ROW_RE = re.compile(r'<row\b([^>]*?)(/>|>(.*?)</row>)', re.S)
_CELL_RE = re.compile(r'<c\b([^>]*?)(/>|>(.*?)</c>)', re.S)


def _attr(s: str, name: str):
    m = re.search(r'\b' + name + r'="([^"]*)"', s)
    return m.group(1) if m else None


def _make_cell(ref: str, style: str | None, value, sst_id) -> str:
    st = f' s="{style}"' if style else ''
    if value is None or value == '':
        return f'<c r="{ref}"{st}/>'
    if isinstance(value, bool):
        return f'<c r="{ref}"{st} t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float)):
        v = repr(value) if isinstance(value, float) else str(value)
        if isinstance(value, float) and v.endswith('.0'):
            v = v[:-2]
        return f'<c r="{ref}"{st}><v>{v}</v></c>'
    sid = sst_id(str(value))
    return f'<c r="{ref}"{st} t="s"><v>{sid}</v></c>'


def _patch_sheet(xml: str, cells: dict[tuple[int, int], object], sst_id) -> str:
    """cells: {(row0, col0): value}

    row0/col0 are 0-based indices into the DATA area, so the Excel row number
    is row0 + 2 (row 1 holds the header).
    """
    # group by Excel row number
    by_row: dict[int, dict[int, object]] = {}
    for (r0, c0), v in cells.items():
        by_row.setdefault(r0 + 2, {})[c0] = v

    m = re.search(r'<sheetData\b[^>]*?(/>|>)', xml)
    if not m:
        raise RuntimeError('malformed sheetData element')
    if m.group(1) == '/>':
        head = xml[:m.start()] + '<sheetData>'
        body = ''
        tail = '</sheetData>' + xml[m.end():]
    else:
        end = xml.index('</sheetData>', m.end())
        head = xml[:m.end()]
        body = xml[m.end():end]
        tail = xml[end:]

    pieces = []          # ordered mix of literal text and row elements
    pos = 0
    existing = set()
    for rm in _ROW_RE.finditer(body):
        pieces.append((body[pos:rm.start()], None, None))
        rnum = int(_attr(rm.group(1), 'r') or 0)
        existing.add(rnum)
        pieces.append((None, rnum, rm.group(0)))
        pos = rm.end()
    pieces.append((body[pos:], None, None))

    style_hint: dict[int, str] = {}

    def patch_row(rowxml: str, rnum: int, edits: dict[int, object]) -> str:
        rm = _ROW_RE.match(rowxml)
        attrs, closing, inner = rm.group(1), rm.group(2), rm.group(3) or ''
        cells_out = []
        seen = {}
        p = 0
        for cm in _CELL_RE.finditer(inner):
            cells_out.append((None, inner[p:cm.start()]))
            ref = _attr(cm.group(1), 'r') or ''
            letters = re.match(r'([A-Z]+)', ref)
            ci = col_index(letters.group(1)) if letters else -1
            st = _attr(cm.group(1), 's')
            if st:
                style_hint.setdefault(ci, st)
            if ci in edits:
                cells_out.append((ci, _make_cell(ref, st, edits[ci], sst_id)))
            else:
                cells_out.append((ci, cm.group(0)))
            seen[ci] = True
            p = cm.end()
        trailing = inner[p:]

        for ci, v in sorted(edits.items()):
            if ci in seen:
                continue
            ref = f'{col_letter(ci)}{rnum}'
            cells_out.append((ci, _make_cell(ref, style_hint.get(ci), v, sst_id)))

        indexed = [x for x in cells_out if x[0] is not None]
        indexed.sort(key=lambda x: x[0])
        inner_new = ''.join(x[1] for x in indexed) + trailing
        if closing == '/>':
            attrs = attrs.rstrip()
            return f'<row{attrs}>{inner_new}</row>'
        return f'<row{attrs}>{inner_new}</row>'

    out = []
    handled = set()
    for text, rnum, rowxml in pieces:
        if rnum is None:
            out.append(text)
            continue
        # insert brand-new rows that belong before this one
        for nr in sorted(n for n in by_row if n not in existing and n < rnum
                         and n not in handled):
            out.append(_new_row(nr, by_row[nr], style_hint, sst_id))
            handled.add(nr)
        if rnum in by_row:
            out.append(patch_row(rowxml, rnum, by_row[rnum]))
            handled.add(rnum)
        else:
            out.append(rowxml)
    for nr in sorted(n for n in by_row if n not in handled):
        out.append(_new_row(nr, by_row[nr], style_hint, sst_id))

    return head + ''.join(out) + tail


def _new_row(rnum: int, edits: dict[int, object], style_hint, sst_id) -> str:
    cs = ''.join(_make_cell(f'{col_letter(ci)}{rnum}', style_hint.get(ci), v, sst_id)
                 for ci, v in sorted(edits.items()))
    return f'<row r="{rnum}">{cs}</row>'
