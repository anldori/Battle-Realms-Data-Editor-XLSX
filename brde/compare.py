"""
brde.compare - Compare the open workbook against another Battle Realms.xlsx

The comparison matches sheets by name, columns by header text and rows by their
primary key (the first column) whenever that key is unique. Falling back to
position only when there is no usable key means inserting one record near the top
of a sheet no longer reports every row below it as changed.
"""
from __future__ import annotations

import csv
import os

import openpyxl

from PyQt6.QtCore import (QAbstractTableModel, QModelIndex, Qt, QTimer,
                          pyqtSignal)
from PyQt6.QtGui import QBrush, QColor
from PyQt6.QtWidgets import (QAbstractItemView, QComboBox, QDialog, QFileDialog,
                             QHBoxLayout, QHeaderView, QLabel, QLineEdit,
                             QMessageBox, QPushButton, QTableView, QVBoxLayout)

CHANGED = 'changed'
ONLY_MINE = 'only_mine'
ONLY_THEIRS = 'only_theirs'

KIND_LABEL = {
    CHANGED: 'changed',
    ONLY_MINE: 'only in this file',
    ONLY_THEIRS: 'only in other file',
}

ROW_BG = {
    CHANGED: QColor(232, 221, 250),
    ONLY_MINE: QColor(214, 240, 220),
    ONLY_THEIRS: QColor(255, 226, 214),
}


# --------------------------------------------------------------------- engine
def _norm(v):
    """Normalise a cell value so trivial encoding differences do not count."""
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def load_values(path, progress=None):
    """Read a workbook into {sheet_name: (headers, rows)} without any inference."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    names = wb.sheetnames
    for i, name in enumerate(names):
        if progress:
            progress(i, len(names), name)
        rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
        while rows and all(v is None for v in rows[-1]):
            rows.pop()
        headers = [('' if h is None else str(h)) for h in rows[0]] if rows else []
        out[name] = (headers, rows[1:] if rows else [])
    wb.close()
    return out


def _key_map(rows):
    """{primary key -> row index} if the first column is a usable unique key."""
    keys = {}
    for i, r in enumerate(rows):
        if not r:
            return None
        k = _norm(r[0])
        if k is None or not isinstance(k, int):
            return None
        if k in keys:
            return None                    # not unique, fall back to position
        keys[k] = i
    return keys


class Diff:
    __slots__ = ('sheet', 'row', 'col', 'header', 'mine', 'theirs', 'kind', 'key')

    def __init__(self, sheet, row, col, header, mine, theirs, kind, key=None):
        self.sheet = sheet
        self.row = row                     # 0-based index into the data area
        self.col = col
        self.header = header
        self.mine = mine
        self.theirs = theirs
        self.kind = kind
        self.key = key


class CompareResult:
    def __init__(self, other_path):
        self.other_path = other_path
        self.diffs: list[Diff] = []
        self.sheets_only_mine: list[str] = []
        self.sheets_only_theirs: list[str] = []
        self.cols_only_mine: list[tuple[str, str]] = []
        self.cols_only_theirs: list[tuple[str, str]] = []

    def per_sheet_counts(self):
        c = {}
        for d in self.diffs:
            c[d.sheet] = c.get(d.sheet, 0) + 1
        return dict(sorted(c.items(), key=lambda kv: -kv[1]))

    def cell_map(self):
        """{(sheet, row, col) -> (mine, theirs)} for highlighting in the grid."""
        return {(d.sheet, d.row, d.col): (d.mine, d.theirs)
                for d in self.diffs if d.col >= 0}


def compare(book, other_path, progress=None) -> CompareResult:
    """Compare the in-memory workbook `book` against the file at `other_path`."""
    res = CompareResult(other_path)
    theirs_all = load_values(other_path, progress)

    mine_names = [n for n in book.sheet_order]
    theirs_names = list(theirs_all.keys())
    res.sheets_only_mine = [n for n in mine_names if n not in theirs_all]
    res.sheets_only_theirs = [n for n in theirs_names if n not in book.sheets]

    common = [n for n in mine_names if n in theirs_all]
    total = len(common)

    for si, name in enumerate(common):
        if progress:
            progress(si, total, name)
        sd = book.sheets[name]
        h_mine = sd.headers
        h_theirs, rows_theirs = theirs_all[name]

        # match columns by header text, not by position
        idx_theirs = {}
        for i, h in enumerate(h_theirs):
            if h and h not in idx_theirs:
                idx_theirs[h] = i
        pairs = []
        for cm, h in enumerate(h_mine):
            if h and h in idx_theirs:
                pairs.append((cm, idx_theirs[h], h))
        res.cols_only_mine += [(name, h) for h in h_mine
                               if h and h not in idx_theirs]
        seen_mine = {h for h in h_mine if h}
        res.cols_only_theirs += [(name, h) for h in h_theirs
                                 if h and h not in seen_mine]

        rows_mine = sd.rows
        km, kt = _key_map(rows_mine), _key_map(rows_theirs)

        if km is not None and kt is not None:
            shared = [k for k in km if k in kt]
            for k in shared:
                _cmp_row(res, book, name, km[k], rows_theirs[kt[k]], pairs, k)
            for k in km:
                if k not in kt:
                    res.diffs.append(Diff(name, km[k], 0, h_mine[0] if h_mine else '',
                                          k, None, ONLY_MINE, k))
            for k in kt:
                if k not in km:
                    res.diffs.append(Diff(name, -1, 0, h_mine[0] if h_mine else '',
                                          None, k, ONLY_THEIRS, k))
        else:
            n = max(len(rows_mine), len(rows_theirs))
            for i in range(n):
                if i >= len(rows_theirs):
                    res.diffs.append(Diff(name, i, -1, '', f'row {i + 2}', None,
                                          ONLY_MINE))
                elif i >= len(rows_mine):
                    res.diffs.append(Diff(name, -1, -1, '', None, f'row {i + 2}',
                                          ONLY_THEIRS))
                else:
                    _cmp_row(res, book, name, i, rows_theirs[i], pairs, None)
    return res


def _cmp_row(res, book, sheet, row_mine, row_theirs, pairs, key):
    for cm, ct, header in pairs:
        a = _norm(book.value(sheet, row_mine, cm))
        b = _norm(row_theirs[ct] if ct < len(row_theirs) else None)
        if a != b:
            res.diffs.append(Diff(sheet, row_mine, cm, header, a, b, CHANGED, key))


# ---------------------------------------------------------------- table model
class DiffModel(QAbstractTableModel):
    HEADERS = ['Sheet', 'Row', 'Key', 'Column', 'This file', 'Other file', 'Status']

    def __init__(self, book, diffs, parent=None):
        super().__init__(parent)
        self.book = book
        self.all = diffs
        self.rows = list(diffs)

    def set_rows(self, rows):
        self.beginResetModel()
        self.rows = rows
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else len(self.HEADERS)

    def headerData(self, s, o, role=Qt.ItemDataRole.DisplayRole):
        if o == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self.HEADERS[s]
        if o == Qt.Orientation.Vertical and role == Qt.ItemDataRole.DisplayRole:
            return str(s + 1)
        return None

    def _label(self, d, value):
        """Show 'code - DESCRIPTION' for enum columns, like the main grid does."""
        if value is None:
            return ''
        if d.col >= 0 and isinstance(value, int):
            tbl = self.book.enum_for(d.sheet, d.col)
            if tbl:
                desc = tbl.code2desc.get(value)
                if desc:
                    return f'{value} - {desc}'
        return str(value)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        d = self.rows[index.row()]
        c = index.column()
        if role == Qt.ItemDataRole.DisplayRole:
            if c == 0:
                return d.sheet
            if c == 1:
                return str(d.row + 2) if d.row >= 0 else '-'
            if c == 2:
                return '' if d.key is None else str(d.key)
            if c == 3:
                return d.header
            if c == 4:
                return self._label(d, d.mine)
            if c == 5:
                return self._label(d, d.theirs)
            if c == 6:
                return KIND_LABEL[d.kind]
        if role == Qt.ItemDataRole.BackgroundRole:
            return QBrush(ROW_BG[d.kind])
        if role == Qt.ItemDataRole.ToolTipRole:
            return (f'{d.sheet}  ·  row {d.row + 2 if d.row >= 0 else "-"}  ·  '
                    f'{d.header}\nThis file: {d.mine}\nOther file: {d.theirs}')
        return None

    def diff_at(self, row):
        return self.rows[row]


# -------------------------------------------------------------------- dialog
class CompareDialog(QDialog):
    """Difference report. Non-modal, so the grid stays usable behind it."""

    jumpRequested = pyqtSignal(str, int, int)          # sheet, row, col
    applyRequested = pyqtSignal(list)                  # [(sheet,row,col,value)]
    highlightToggled = pyqtSignal(bool)

    # Column widths. The three narrow columns hold a number or one short word
    # and keep a fixed pixel width; the four text columns share whatever the
    # window is not otherwise using, in these proportions.
    FIXED_W = {1: 60, 2: 60, 6: 120}                   # Row, Key, Status
    FLEX_W = {0: 21, 3: 19, 4: 21, 5: 21}              # Sheet, Column, both files
    MIN_FLEX = 80

    def __init__(self, book, result: CompareResult, parent=None):
        super().__init__(parent)
        self.book = book
        self.result = result
        self._applying_widths = False
        self._user_sized = False
        # Parented, so a dialog closed before it fires takes the timer with it.
        self._settle_timer = QTimer(self)
        self._settle_timer.setSingleShot(True)
        self._settle_timer.timeout.connect(self._settle_columns)
        self.setWindowTitle('Compare - differences')
        self.resize(1100, 620)
        self.setSizeGripEnabled(True)

        v = QVBoxLayout(self)

        # ---- summary
        n = len(result.diffs)
        changed = sum(1 for d in result.diffs if d.kind == CHANGED)
        only_m = sum(1 for d in result.diffs if d.kind == ONLY_MINE)
        only_t = sum(1 for d in result.diffs if d.kind == ONLY_THEIRS)
        head = QLabel(
            f'<b>{n} differences</b> &nbsp; ({changed} changed cells, '
            f'{only_m} records only here, {only_t} records only in the other file)'
            f'<br>Comparing against: <code>{result.other_path}</code>')
        head.setWordWrap(True)
        v.addWidget(head)

        extra = []
        if result.sheets_only_mine:
            extra.append(f'Sheets only in this file: {", ".join(result.sheets_only_mine)}')
        if result.sheets_only_theirs:
            extra.append(f'Sheets only in the other file: {", ".join(result.sheets_only_theirs)}')
        if result.cols_only_mine:
            extra.append('Columns only in this file: ' +
                         ', '.join(f'{s}.{c}' for s, c in result.cols_only_mine[:12]))
        if result.cols_only_theirs:
            extra.append('Columns only in the other file: ' +
                         ', '.join(f'{s}.{c}' for s, c in result.cols_only_theirs[:12]))
        if extra:
            lab = QLabel('\n'.join(extra))
            lab.setWordWrap(True)
            lab.setStyleSheet('color:#5a6b85;')
            v.addWidget(lab)

        # ---- filters
        bar = QHBoxLayout()
        bar.addWidget(QLabel('Sheet:'))
        self.cb_sheet = QComboBox()
        self.cb_sheet.setMinimumWidth(220)
        counts = result.per_sheet_counts()
        self.cb_sheet.addItem(f'All sheets ({n})', '')
        for s, cnt in counts.items():
            self.cb_sheet.addItem(f'{s}  ({cnt})', s)
        self.cb_sheet.currentIndexChanged.connect(self._refilter)
        bar.addWidget(self.cb_sheet)

        self.lbl_kind = QLabel('   Status:')
        bar.addWidget(self.lbl_kind)
        self.cb_kind = QComboBox()
        self.cb_kind.addItem('All', '')
        self.cb_kind.addItem('Changed cells', CHANGED)
        self.cb_kind.addItem('Only in this file', ONLY_MINE)
        self.cb_kind.addItem('Only in other file', ONLY_THEIRS)
        self.cb_kind.currentIndexChanged.connect(self._refilter)
        bar.addWidget(self.cb_kind)

        bar.addWidget(QLabel('   Find:'))
        self.ed = QLineEdit()
        self.ed.setPlaceholderText('column name, value, description...')
        self.ed.textChanged.connect(self._refilter)
        bar.addWidget(self.ed, 1)
        v.addLayout(bar)

        # ---- table
        self.model = DiffModel(book, result.diffs, self)
        self.tbl = QTableView()
        self.tbl.setModel(self.model)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.verticalHeader().setDefaultSectionSize(22)
        self.tbl.doubleClicked.connect(self._on_double_click)
        hh = self.tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        v.addWidget(self.tbl, 1)
        self._size_columns()

        # "Status" only says something when the report holds more than one kind
        # of difference. Two files with the same records differ in cell values
        # alone, and then the column reads "changed" on every row and the filter
        # has one usable entry, so both are hidden.
        if len({d.kind for d in result.diffs}) < 2:
            self.tbl.setColumnHidden(DiffModel.HEADERS.index('Status'), True)
            self.lbl_kind.hide()
            self.cb_kind.hide()
            self._fit_columns()

        # Connected last: hiding a column and the initial sizing above both
        # emit sectionResized, and neither is the user dragging a divider.
        hh.sectionResized.connect(self._on_section_resized)

        # ---- buttons
        row = QHBoxLayout()
        self.chk_hl = QPushButton('Highlight in grid')
        self.chk_hl.setCheckable(True)
        self.chk_hl.setChecked(True)
        self.chk_hl.toggled.connect(self.highlightToggled.emit)
        row.addWidget(self.chk_hl)

        b_jump = QPushButton('Go to cell')
        b_jump.clicked.connect(self._jump_selected)
        row.addWidget(b_jump)

        b_take = QPushButton('Take other value')
        b_take.setToolTip('Copy the other file\'s value into this file as a normal,\n'
                          'undoable edit. Save with Ctrl+S to write it out.')
        b_take.clicked.connect(self._take_selected)
        # The report itself is worth having against a read-only file - seeing
        # what an old .dat holds that the workbook does not is most of the point
        # of opening one - but there is nowhere to copy a value into.
        if self.book.read_only:
            b_take.setEnabled(False)
            b_take.setToolTip('This file was opened read-only.')
        row.addWidget(b_take)

        b_csv = QPushButton('Export to CSV...')
        b_csv.clicked.connect(self._export)
        row.addWidget(b_csv)

        row.addStretch(1)
        b_close = QPushButton('Close')
        b_close.clicked.connect(self.close)
        row.addWidget(b_close)
        v.addLayout(row)

        self.lbl_hint = QLabel('Double-click a row to jump to that cell in the grid.')
        self.lbl_hint.setStyleSheet('color:#5a6b85;')
        v.addWidget(self.lbl_hint)

    # ------------------------------------------------------------ geometry
    def _size_columns(self):
        for c, w in self.FIXED_W.items():
            self.tbl.setColumnWidth(c, w)
        self._fit_columns()

    def _visible_cols(self):
        return [c for c in range(len(DiffModel.HEADERS))
                if not self.tbl.isColumnHidden(c)]

    def _fit_columns(self):
        """Hand the text columns the width the fixed ones leave over.

        Sized twice: once now, so the window never paints a half-filled table,
        and once after the event loop has caught up. The second pass is what
        makes it exact - widening the columns can push the vertical scrollbar
        out of the table, and the viewport is still the narrower size when
        `resizeEvent` measures it, which leaves a scrollbar-wide empty strip.
        """
        if self._user_sized:
            return
        self._apply_widths()
        self._settle_timer.start(0)

    def _settle_columns(self):
        if self._user_sized:
            return
        if sum(self.tbl.columnWidth(c) for c in self._visible_cols()) \
                != self.tbl.viewport().width():
            self._apply_widths()

    def _apply_widths(self):
        """One sizing pass over the flexible columns."""
        avail = self.tbl.viewport().width()
        for c, w in self.FIXED_W.items():
            if not self.tbl.isColumnHidden(c):
                avail -= w
        flex = [c for c in sorted(self.FLEX_W) if not self.tbl.isColumnHidden(c)]
        if not flex or avail <= self.MIN_FLEX * len(flex):
            return
        total = sum(self.FLEX_W[c] for c in flex)
        self._applying_widths = True
        used = 0
        for c in flex[:-1]:
            w = max(self.MIN_FLEX, avail * self.FLEX_W[c] // total)
            self.tbl.setColumnWidth(c, w)
            used += w
        # the last one takes the rounding remainder, so no strip is left over
        self.tbl.setColumnWidth(flex[-1], max(self.MIN_FLEX, avail - used))
        self._applying_widths = False

    def _on_section_resized(self, *_):
        """Once a divider has been dragged, the widths are the user's to keep."""
        if not self._applying_widths:
            self._user_sized = True

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._fit_columns()

    def showEvent(self, e):
        # The viewport has no real width until the layout has run, so the
        # sizing done in __init__ is against a placeholder. Redo it here.
        super().showEvent(e)
        self._fit_columns()

    # ------------------------------------------------------------ filtering
    def _refilter(self, *_):
        sheet = self.cb_sheet.currentData() or ''
        kind = self.cb_kind.currentData() or ''
        needle = self.ed.text().strip().lower()
        rows = []
        for d in self.result.diffs:
            if sheet and d.sheet != sheet:
                continue
            if kind and d.kind != kind:
                continue
            if needle:
                hay = f'{d.sheet} {d.header} {d.mine} {d.theirs} {d.key}'.lower()
                if needle not in hay:
                    tbl = (self.book.enum_for(d.sheet, d.col)
                           if d.col >= 0 else None)
                    hit = False
                    if tbl:
                        for val in (d.mine, d.theirs):
                            if isinstance(val, int):
                                desc = tbl.code2desc.get(val, '')
                                if desc and needle in desc.lower():
                                    hit = True
                    if not hit:
                        continue
            rows.append(d)
        self.model.set_rows(rows)
        self.lbl_hint.setText(f'Showing {len(rows)} of {len(self.result.diffs)} '
                              'differences.  Double-click a row to jump to it.')

    # ------------------------------------------------------------ actions
    def _selected_diffs(self):
        seen, out = set(), []
        for i in self.tbl.selectionModel().selectedIndexes():
            if i.row() not in seen:
                seen.add(i.row())
                out.append(self.model.diff_at(i.row()))
        return out

    def _on_double_click(self, index):
        self._emit_jump(self.model.diff_at(index.row()))

    def _jump_selected(self, *_):
        ds = self._selected_diffs()
        if ds:
            self._emit_jump(ds[0])

    def _emit_jump(self, d):
        if d.row < 0:
            QMessageBox.information(
                self, 'Compare',
                'This record only exists in the other file, so there is no '
                'matching row in the file you have open.')
            return
        self.jumpRequested.emit(d.sheet, d.row, max(d.col, 0))

    def _take_selected(self, *_):
        ds = [d for d in self._selected_diffs()
              if d.kind == CHANGED and d.row >= 0 and d.col >= 0]
        if not ds:
            QMessageBox.information(
                self, 'Compare',
                'Select one or more rows with the status "changed" first.\n\n'
                'Records that exist in only one of the files cannot be copied '
                'this way.')
            return
        r = QMessageBox.question(
            self, 'Take other value',
            f'Copy {len(ds)} value(s) from the other file into this one?\n\n'
            'They become normal edits: undo with Ctrl+Z, write them to disk '
            'with Ctrl+S.',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if r != QMessageBox.StandardButton.Yes:
            return
        self.applyRequested.emit([(d.sheet, d.row, d.col, d.theirs) for d in ds])

    def _export(self, *_):
        base = os.path.splitext(os.path.basename(self.book.path))[0]
        default = os.path.join(os.path.dirname(self.book.path),
                               f'{base}_compare.csv')
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export differences', default,
            'CSV file (*.csv);;Text file (*.txt);;All files (*.*)')
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
                w = csv.writer(fh)
                w.writerow(['# This file', self.book.path])
                w.writerow(['# Other file', self.result.other_path])
                w.writerow([])
                w.writerow(DiffModel.HEADERS)
                for d in self.model.rows:
                    w.writerow([
                        d.sheet,
                        d.row + 2 if d.row >= 0 else '',
                        '' if d.key is None else d.key,
                        d.header,
                        '' if d.mine is None else d.mine,
                        '' if d.theirs is None else d.theirs,
                        KIND_LABEL[d.kind],
                    ])
        except Exception as e:
            QMessageBox.critical(self, 'Compare', f'Could not write the file:\n{e}')
            return
        self.lbl_hint.setText(f'Exported {len(self.model.rows)} rows to {path}')
